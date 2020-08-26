import getpass, getopt
import covscraper
import datetime
import sys, json
import logging
import jinja2

import exchangelib
from exchangelib import Credentials, Configuration, Account
from exchangelib.indexed_properties import EmailAddress
from exchangelib import Message, Mailbox, HTMLBody

import base64

TEMPLATE = """
<b>SID: </b>{{ sid }}<br/>
<b>Name: </b>{{ details.get("firstName") }} {{ details.get("lastName") }}<br/>
<b>Email: </b>{{ details.get("email") }}<br/>
<a href="{{ details.get("url") }}">Student details</a> 
<a href="{{ details.get("engageurl") }}">Attendance data</a><br/>
<b>Percentages</b>
{% if percent == None %}
  <font color="red">No data</font><br/>
{% else %}  
  <ul>
    {% for key, val in percent.items() %}
      <li>
        {% if val < 80 %}
          <font color="red">{{val|round(0)|int}}% {{ key }}</font>
        {% else %}
          {{val|round(0)|int}}% {{ key }}
        {% endif %}
      </li>
    {% endfor %}
  </ul>
{% endif %}

<b>Absolutes</b>
{% if absolute == None %}
  <font color="red">No data</font><br/>
{% else %}  
  <ul>
    {% for key, val in absolute.items() %}
      <li>{{ val }} {{ key }}</li>
    {% endfor %}
  </ul>
{% endif %}<br/>"""


class Outlooker:
    def __init__(self, email, password):
        server = 'outlook.office365.com'
        creds = Credentials(username=email, password=password)
        config = Configuration(server=server, credentials=creds)
        self.account = Account(primary_smtp_address=email, autodiscover=False, 
                        config=config, access_type=exchangelib.DELEGATE)

    def send(self, to, body):
        m = Message(account=self.account,
                    folder=self.account.sent,
                    subject="Personal tutor attendance data",
                    body=HTMLBody(body),
                    to_recipients=[Mailbox(email_address=to)])
        m.send_and_save()

from openpyxl import Workbook, load_workbook
import re

def generate_email_body( students, start, end ):
    jtemplate = jinja2.Environment(loader=jinja2.BaseLoader).from_string(TEMPLATE)

    dateFmt = "%H:%M %d/%m/%Y"
    msgBody = "<h3>Attendance period {} to {}</h3>".format(start.strftime(dateFmt),end.strftime(dateFmt))
    for s in students:
        logger.info( "Process {}".format(s) )
        try:
            details = covscraper.studentapi.get_student_details( session, s )
        except covscraper.studentapi.NoStudent:
            details = {}

        try:      
            engagement = covscraper.studentapi.get_engagement( session, s )
            engagement["sessions"] = [ i for i in engagement["sessions"] if i["start"] >= start and i["end"] <= end ]
            absolute, percent = covscraper.studentapi.get_attendance( engagement )
        except covscraper.studentapi.NoStudent:
            absolute, percent = None, None

        msgBody += jtemplate.render( sid=s, percent=percent, absolute=absolute, details=details )

    return msgBody


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

    usageTxt = "help"
    
    user = None
    pwd = None
    days = 7
    to = None
    file = None
   
    # configure flags
    logger.debug( "Handle CLI args" )

    shortopts = "u:p:d:t:hf:"
    longopts = ["user=","pass=","days=","help","to=","file="]
    try:
        opts, args = getopt.getopt(sys.argv[1:], shortopts, longopts)
    except getopt.GetoptError as e:
        print(e)
        sys.exit(1)
        
    # process flags
    for o, a in opts:
        if o in ("-h", "--help"):
            print(usageText)
            sys.exit(1)
        elif o in ("-l", "--latest"): latest = True
        elif o in ("-u", "--user"): user = a
        elif o in ("-p", "--pass"): pwd = a
        elif o in ("-d", "--days"): days = int(a)
        elif o in ("-t", "--to"): to = a
        elif o in ("-f", "--file"): file = a

    if not user: user = input("username: ")
    if not pwd: pwd = getpass.getpass("password: ")
    
    now = datetime.datetime.now()
    since = now - datetime.timedelta(days=days)
    logger.debug( since )

    # authenticate
    logger.debug( "authenticate" )
    outlook = Outlooker(user+"@coventry.ac.uk", pwd)
    session = covscraper.auth.Authenticator( user, pwd )

    # get student uids
    if to != None:
        outlook.send( to, generate_email_body( ( int(i) for i in args ), since, now ) )


    if file:
        regStaffEmail = re.compile( r"csx[0-9]{3}|[a-z]{2}[0-9]{4}", re.I )

        wb = load_workbook( filename=file )

        ws = wb.get_active_sheet()
        for row in ws.iter_rows( min_row=2 ):
            staffid = row[0].value
            if not staffid: continue
            if not regStaffEmail.match( staffid ): 
                logger.info( "Skipping {} as not a staff email".format(staffid) )
                continue # not a staff email

            logger.info( "Sending to {}".format(staffid) )

            sids = [ i.value for i in row[1:] if i.value != None ]

            outlook.send( staffid+"@coventry.ac.uk", generate_email_body(sids, since, now) )

    sys.exit(0)
